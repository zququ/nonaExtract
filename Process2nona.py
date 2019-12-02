# modify the original file

f = open('sequence.fasta', 'r')
f0 = open('processed_sequence.fasta', 'w')

for lines in f.readlines()[:-1]:
    f0.write(lines)
f.seek(0)
ori_lines = f.readlines()
print(ori_lines)
f0.write(ori_lines[-1] + '@')
f0.close()

f0 = open('processed_sequence.fasta', 'r')
f_countAA = open('deSample_output.txt', 'w')
f_proName = open('ProName_output.txt', 'w')

for each_line in f0:
    print(each_line)
    Name_Len_count = 0
    if each_line[0]=='>':
        for each_str in each_line:
            Name_Len_count += 1
            if each_str == '[':
                Name_Len = Name_Len_count
                Name_Len_count = 0

        f_countAA.write('@')
        f_proName.write(each_line[1:Name_Len-2] + '\n')
    else:
        f_countAA.write(each_line)

f.close()
f0.close()
f_countAA.close()
f_proName.close()

f2 = open('deSample_output.txt')
flist = []
n = 0
for each_str in f2.read():
    n += 1

    if each_str == '\n':
        flist.append('')
    else:
        flist.append(each_str)

f3 = open('deSample_output_deenter.txt', 'w+')
ls = ''.join(flist)
f3.write(ls)

f2.close()
f3.close()


f4 = open('deSample_output_deenter.txt', 'r')
f5 = open('process2nona_output.txt', 'w+')
lst = []
n = 0

for each in f4:
    lst.extend(each)
    #print(lst)

for i in range(len(lst)-9):
    #print(lst[i:i+9])   FOR TEST
    ls_out = ''.join(lst[i:i+9])
    print(ls_out)
    output = str(ls_out)
    if not '@' in output:
        f5.write(output + '\n')

f4.close()
f5.close()


f6 = open('deSample_output_deenter.txt', 'r')
f7 = open('count_result.txt', 'w+')
list1 = []
n = 0

for each_str in f6.read():
    n += 1
    if each_str == '@':
        if n-9 > 0:
            print(n-9)  #n-1
            n_str = str(n-9)
            f7.write(n_str + '\n')
        n = 0
    else:
        continue

f6.close()
f7.close()

f8 = open('ProName_output.txt', 'r')
f9 = open('proName_and_number.txt', 'w')
f10 = open('count_result.txt', 'r')

n_new = 0
line_number = 0
limit_line = f10.readlines()[line_number]
f10 = open('count_result.txt', 'r')
limit_number = len(f10.readlines())
#print(limit_number)
f8.close()
f9.close()
f10.close()
#print(ProName[line_number][:-1])

for i in range(limit_number):
    f8 = open('ProName_output.txt', 'r')
    f9 = open('proName_and_number.txt', 'a')
    f10 = open('count_result.txt')
    n_new = 0
    limit_line = f10.readlines()[line_number]
    print(limit_line)
    while n_new < int(limit_line):
        f8 = open('ProName_output.txt', 'r')
        f9 = open('proName_and_number.txt', 'a')
        f10 = open('count_result.txt', 'r')
        ProName = f8.readlines()
        f9.write(ProName[line_number][:-1] + '_' + str(n_new+1) + '\n')
        f9.close()
        n_new += 1
        f8.close()
        f9.close()
        f10.close()
    line_number += 1
    f8.close()
    f9.close()
    f10.close()


f8.close()
f9.close()
f10.close()


dict1 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }

dict2 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }


dict9 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }

f5 = open('process2nona_output.txt', 'r')
f6 = open('value.txt', 'w')
nameline = f5.readlines()
total_value = 0
for each_line in nameline:
    value1 = 0
    for each_str1 in each_line[0]:    # modify the motif position
        for each_str2 in each_line[1]:
            for each_str9 in each_line[8]:
                value1 = dict1[each_str1]
                value2 = dict2[each_str2]
                value9 = dict9[each_str9]

                f6 = open('value.txt', 'a')
                f6.write(str(value1 + value2 + value9) + '\n')
                f6.close()

f5.close()
f6.close()


import openpyxl

f9 = open('proName_and_number.txt', 'r')
f5 = open('process2nona_output.txt', 'r')
f12 = open('value.txt', 'r')

wb = openpyxl.Workbook()
ws = wb.active
n = 0
limit_number = len(f9.readlines())
f9.seek(0)

for line in range(limit_number):
    f9 = open('proName_and_number.txt', 'r')
    f5 = open('process2nona_output.txt', 'r')
    f12 = open('value.txt', 'r')

    name_inf = f9.readlines()[n]
    pep_inf = f5.readlines()[n]
    value_inf = f12.readlines()[n]
    n += 1
    name1 = 'A' + str(n)
    name2 = 'B' + str(n)
    name3 = 'C' + str(n)
    ws[name1] = name_inf
    ws[name2] = pep_inf
    ws[name3] = value_inf

    f9.close()
    f5.close()
    f12.close()


wb.save("test.xlsx")

f9.close()
f5.close()

