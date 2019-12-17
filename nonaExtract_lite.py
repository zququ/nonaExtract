#
#  Created by Zehui Qu on 2019/12/02.
#  Copyright © 2019 zququ. All rights reserved.
#
# nonaExtract
# Used to process fasta format AA sequence into nona-peptide and scoring

# Instruction:

# 1. Install python.
# 2. Install the dependence, openpyxl. Maybe, just with command line: pip install openpyxl.
# 3. Make a fasta file with fasta formatted sequence, maybe just download from the NCBI protein database.
# 4. Run nonaExtract.py.

# To use this easy script, please cite:

# Structure and Peptidome of the Bat MHC Class I Molecule Reveal a Novel Mechanism Leading to High-Affinity Peptide Binding Zehui Qu, Zibin Li, Lizhen Ma, Xiaohui Wei, Lijie Zhang, Ruiying Liang, Geng Meng, Nianzhi Zhang and Chun Xia J Immunol May 10, 2019, ji1900001; DOI: https://doi.org/10.4049/jimmunol.1900001

# Any further question, feel free to contact the author, Zehui Qu: qzh813@gmail.com

# modify the original file
f = open('sequence.fasta', 'r')
f0 = open('processed_sequence.fasta', 'w')

for lines in f.readlines()[:-1]:
    f0.write(lines)
f.seek(0)
ori_lines = f.readlines()
print(ori_lines)
f0.write(ori_lines[-1] + '@')
f0.close()

# Process The sequence.fasta to nonapeptide
f0 = open('processed_sequence.fasta', 'r')
f_countAA = open('deSample_output.txt', 'w')
f_proName = open('ProName_output.txt', 'w')

for each_line in f0:
    print(each_line)
    Name_Len_count = 0
    if each_line[0]=='>':
        for each_str in each_line:
            Name_Len_count += 1
            if each_str == '[':
                Name_Len = Name_Len_count
                Name_Len_count = 0

        f_countAA.write('@')
        f_proName.write(each_line[1:Name_Len-2] + '\n')
    else:
        f_countAA.write(each_line)

f.close()
f0.close()
f_countAA.close()
f_proName.close()

f2 = open('deSample_output.txt')
flist = []
n = 0
for each_str in f2.read():
    n += 1

    if each_str == '\n':
        flist.append('')
    else:
        flist.append(each_str)

f3 = open('deSample_output_deenter.txt', 'w+')
ls = ''.join(flist)
f3.write(ls)

f2.close()
f3.close()


f4 = open('deSample_output_deenter.txt', 'r')
f5 = open('process2nona_output.txt', 'w+')
lst = []
n = 0

for each in f4:
    lst.extend(each)
    #print(lst)

for i in range(len(lst)-9):
    #print(lst[i:i+9])   FOR TEST
    ls_out = ''.join(lst[i:i+9])
    print(ls_out)
    output = str(ls_out)
    if not '@' in output:
        f5.write(output + '\n')

f4.close()
f5.close()

# count for sequence, locate the sequence number

f6 = open('deSample_output_deenter.txt', 'r')
f7 = open('count_result.txt', 'w+')
list1 = []
n = 0

for each_str in f6.read():
    n += 1
    if each_str == '@':
        if n-9 > 0:
            print(n-9)  #n-1
            n_str = str(n-9)
            f7.write(n_str + '\n')
        n = 0
    else:
        continue

f6.close()
f7.close()

f8 = open('ProName_output.txt', 'r')
f9 = open('proName_and_number.txt', 'w')
f10 = open('count_result.txt', 'r')

n_new = 0
line_number = 0
limit_line = f10.readlines()[line_number]
f10 = open('count_result.txt', 'r')
limit_number = len(f10.readlines())
#print(limit_number)
f8.close()
f9.close()
f10.close()
#print(ProName[line_number][:-1])

# link the Protein Name and the sequence number

for i in range(limit_number):
    f8 = open('ProName_output.txt', 'r')
    f9 = open('proName_and_number.txt', 'a')
    f10 = open('count_result.txt')
    n_new = 0
    limit_line = f10.readlines()[line_number]
    print(limit_line)
    while n_new < int(limit_line):
        f8 = open('ProName_output.txt', 'r')
        f9 = open('proName_and_number.txt', 'a')
        f10 = open('count_result.txt', 'r')
        ProName = f8.readlines()
        f9.write(ProName[line_number][:-1] + '_' + str(n_new+1) + '\n')
        f9.close()
        n_new += 1
        f8.close()
        f9.close()
        f10.close()
    line_number += 1
    f8.close()
    f9.close()
    f10.close()

f8.close()
f9.close()
f10.close()

# calculate the score, change the value


dict1 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }

dict2 = {'A': 1,
        'C': 1,
        'D': 2,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 2,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 2,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 2
        }

dict3 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }


dict9 = {'A': 1,
        'C': 1,
        'D': 1,
        'E': 1,
        'F': 1,
        'G': 1,
        'H': 1,
        'I': 1,
        'K': 1,
        'L': 1,
        'M': 1,
        'N': 1,
        'P': 1,
        'Q': 1,
        'R': 1,
        'S': 1,
        'T': 1,
        'V': 1,
        'W': 1,
        'Y': 1
        }


# not used usullY

dict4 = {'A': 0,
        'C': 0,
        'D': 0,
        'E': 0,
        'F': 0,
        'G': 0,
        'H': 0,
        'I': 0,
        'K': 0,
        'L': 0,
        'M': 0,
        'N': 0,
        'P': 0,
        'Q': 0,
        'R': 0,
        'S': 0,
        'T': 0,
        'V': 0,
        'W': 0,
        'Y': 0
        }


dict5 = {'A': 0,
        'C': 0,
        'D': 0,
        'E': 0,
        'F': 0,
        'G': 0,
        'H': 0,
        'I': 0,
        'K': 0,
        'L': 0,
        'M': 0,
        'N': 0,
        'P': 0,
        'Q': 0,
        'R': 0,
        'S': 0,
        'T': 0,
        'V': 0,
        'W': 0,
        'Y': 0
        }

dict6 = {'A': 0,
        'C': 0,
        'D': 0,
        'E': 0,
        'F': 0,
        'G': 0,
        'H': 0,
        'I': 0,
        'K': 0,
        'L': 0,
        'M': 0,
        'N': 0,
        'P': 0,
        'Q': 0,
        'R': 0,
        'S': 0,
        'T': 0,
        'V': 0,
        'W': 0,
        'Y': 0
        }


dict7 = {'A': 0,
        'C': 0,
        'D': 0,
        'E': 0,
        'F': 0,
        'G': 0,
        'H': 0,
        'I': 0,
        'K': 0,
        'L': 0,
        'M': 0,
        'N': 0,
        'P': 0,
        'Q': 0,
        'R': 0,
        'S': 0,
        'T': 0,
        'V': 0,
        'W': 0,
        'Y': 0
        }


dict8 = {'A': 0,
        'C': 0,
        'D': 0,
        'E': 0,
        'F': 0,
        'G': 0,
        'H': 0,
        'I': 0,
        'K': 0,
        'L': 0,
        'M': 0,
        'N': 0,
        'P': 0,
        'Q': 0,
        'R': 0,
        'S': 0,
        'T': 0,
        'V': 0,
        'W': 0,
        'Y': 0
        }

f5 = open('process2nona_output.txt', 'r')
f6 = open('value.txt', 'w')
nameline = f5.readlines()
total_value = 0
for each_line in nameline:
    value1 = 0
    value2 = 0
    value3 = 0
    value4 = 0
    value5 = 0
    value6 = 0
    value7 = 0
    value8 = 0
    value9 = 0
    for each_str1 in each_line[0]:    # modify the motif position  
        for each_str2 in each_line[1]:
            for each_str9 in each_line[8]:
                for each_str3 in each_line[2]:
                    for each_str4 in each_line[3]:
                        for each_str5 in each_line[4]:
                            for each_str6 in each_line[5]:
                                for each_str7 in each_line[6]:
                                    for each_str8 in each_line[7]:
                                        value1 = dict1[each_str1]
                                        value2 = dict2[each_str2]
                                        value3 = dict3[each_str3]
                                        value4 = dict4[each_str4]
                                        value5 = dict5[each_str5]
                                        value6 = dict6[each_str6]
                                        value7 = dict7[each_str7]
                                        value8 = dict8[each_str8]
                                        value9 = dict9[each_str9]
                                        f6 = open('value.txt', 'a')
                                        f6.write(str(value1 + value2 + value9 + value3 + value4 + value5 + value6 + value7 + value8) + '\n')
                                        f6.close()

f5.close()
f6.close()


import openpyxl

f9 = open('proName_and_number.txt', 'r')

wb = openpyxl.Workbook()
ws = wb.active
n = 0
limit_number = len(f9.readlines())

for line in range(limit_number):
    f9 = open('proName_and_number.txt', 'r')
    f5 = open('process2nona_output.txt', 'r')
    f12 = open('value.txt', 'r')

    name_inf = f9.readlines()[n]
    pep_inf = f5.readlines()[n]
    value_inf = f12.readlines()[n]
    n += 1
    name1 = 'A' + str(n)
    name2 = 'B' + str(n)
    name3 = 'C' + str(n)
    ws[name1] = name_inf
    ws[name2] = pep_inf
    ws[name3] = value_inf

    f9.close()
    f5.close()
    f12.close()


wb.save("output.xlsx")

f9.close()
f5.close()
f12.close()

import os, sys

os.remove('count_result.txt')
os.remove('deSample_output_deenter.txt')
os.remove('deSample_output.txt')
# os.remove('process2nona_output.txt')
os.remove('processed_sequence.fasta')
# os.remove('proName_and_number.txt')
os.remove('ProName_output.txt')
# os.remove('value.txt')
#, 'deSample_output_deenter.txt', 'deSample_output.txt', 'process2nona_output.txt', 'processed_sequence.fasta', 'proName_and_number.txt', 'ProName_output.txt', 'value.txt')

